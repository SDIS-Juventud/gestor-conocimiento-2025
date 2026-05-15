# -*- coding: utf-8 -*-
"""Genera el Excel ejes/Políticas/Configuracion_politicas_bienestar.xlsx
con dos hojas editables:

  - Exclusiones: productos del Excel oficial que NO deben aparecer en la
    sección "Reporte a políticas" de Bienestar (porque corresponden a
    otros ejes).
  - Reglas_cruce: explicaciones de cómo se reporta cada producto que no
    tiene código de actividad SIRBE propio (se reportan por cruce de
    variables de caracterización en la ficha SIRBE).

Se corre UNA sola vez para sembrar el Excel. Después Carolina lo edita
a mano y el script generar_eje_bienestar.py lee los valores actualizados.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
RUTA = BASE / "ejes" / "Políticas" / "Configuracion_politicas_bienestar.xlsx"

EXCLUSIONES = [
    # (Tema contiene, Producto contiene, Notas)
    ("Comité Intersectorial de Salud", "Orientación socio-ocupacional",
     "Este producto corresponde al eje Inclusión social y productiva, no a Bienestar."),
    ("SIDFAC", "",
     "El Sistema Distrital de Formación Artística y Cultural corresponde al eje Cultura."),
    ("Formación Artística y Cultural", "",
     "Producto del eje Cultura, no de Bienestar."),
]

REGLAS_CRUCE = [
    # (Patrón en tema/producto, Cómo se reporta)
    ("LGBTI",
     "Ficha SIRBE casillas E (identidad de género), F (identidad transgénero) y G (orientación sexual). "
     "El equipo de la DADE filtra a los y las jóvenes que se autorreconocen en sectores LGBTI."),
    ("Negra",
     "Ficha SIRBE casilla AA (grupo étnico). El equipo de la DADE toma a los y las jóvenes que reportaron "
     "autorreconocimiento étnico."),
    ("Afrocolombiana",
     "Ficha SIRBE casilla AA (grupo étnico). El equipo de la DADE toma a los y las jóvenes que reportaron "
     "autorreconocimiento étnico."),
    ("Palenquera",
     "Ficha SIRBE casilla AA (grupo étnico, opción Palenquero de San Basilio)."),
    ("Indígena",
     "Ficha SIRBE casillas AA (grupo étnico) y AB (grupo indígena específico)."),
    ("Raizal",
     "Ficha SIRBE casilla AA (grupo étnico, opción Raizal)."),
    ("Rrom",
     "Ficha SIRBE casilla AA (grupo étnico, opción Rrom / Gitano)."),
    ("Actividades Sexuales Pagadas",
     "Ficha SIRBE casilla T (Rol Ocupacional, opción Actividades Sexuales pagadas)."),
    ("Víctimas",
     "Cruce interno de la DADE con el Registro Único de Víctimas (RUV) y la ficha SIRBE casilla AC "
     "(víctima de conflicto armado)."),
    ("PDET",
     "Filtro por ubicación territorial: actividades prestadas en la Localidad 20 de Sumapaz o en zonas "
     "rurales priorizadas PDET."),
    ("Reincorporación",
     "Al diligenciar el nombre del curso en SIRBE se agrega el marcador “Reincorporados”, y el equipo "
     "filtra por ese marcador."),
    ("Ruralidad",
     "Ficha SIRBE casilla AH (tipo de predio: rural)."),
    ("Discapacidad",
     "Ficha SIRBE casilla Z (Persona con discapacidad) y tipo de discapacidad."),
    ("Migrantes",
     "Ficha SIRBE casilla A (tipo de documento), opciones de cédula de extranjería, permisos o pasaporte."),
]


def _set_estilo_header(ws, headers):
    fill = PatternFill(start_color="253C5C", end_color="253C5C", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 32


def _set_estilo_filas(ws, n_filas, n_cols):
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for r in range(2, n_filas + 2):
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border
            cell.font = Font(size=10, name="Calibri")


def main():
    wb = openpyxl.Workbook()

    # Hoja 1: Exclusiones
    ws1 = wb.active
    ws1.title = "Exclusiones"
    headers1 = ["Tema contiene", "Producto contiene", "Notas"]
    ws1.append(headers1)
    for fila in EXCLUSIONES:
        ws1.append(list(fila))
    _set_estilo_header(ws1, headers1)
    _set_estilo_filas(ws1, len(EXCLUSIONES), len(headers1))
    anchos1 = [40, 40, 60]
    for i, w in enumerate(anchos1, 1):
        ws1.column_dimensions[chr(64 + i)].width = w
    ws1.freeze_panes = "A2"

    # Hoja 2: Reglas de cruce
    ws2 = wb.create_sheet("Reglas_cruce")
    headers2 = ["Patrón en tema o producto", "Cómo se reporta"]
    ws2.append(headers2)
    for fila in REGLAS_CRUCE:
        ws2.append(list(fila))
    _set_estilo_header(ws2, headers2)
    _set_estilo_filas(ws2, len(REGLAS_CRUCE), len(headers2))
    anchos2 = [28, 90]
    for i, w in enumerate(anchos2, 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A2"

    # Hoja 3: Instrucciones (referencia para Carolina)
    ws3 = wb.create_sheet("Instrucciones")
    ws3.append(["Configuración de la sección 'Reporte a políticas' del eje Bienestar"])
    ws3.append([""])
    ws3.append(["Hoja 'Exclusiones':"])
    ws3.append(["  Productos del Excel oficial de Felipe que NO deben aparecer en"])
    ws3.append(["  la sección Reporte a políticas de Bienestar (porque corresponden"])
    ws3.append(["  a otros ejes: Inclusión, Cultura, Liderazgo, etc.)."])
    ws3.append([""])
    ws3.append(["  Para excluir un producto: agregar una fila con substring del tema"])
    ws3.append(["  y opcionalmente substring del nombre del producto. La comparación"])
    ws3.append(["  ignora tildes y mayúsculas."])
    ws3.append([""])
    ws3.append(["Hoja 'Reglas_cruce':"])
    ws3.append(["  Explica cómo se reporta cada política o plan que NO tiene un código"])
    ws3.append(["  SIRBE propio. El reporte se hace cruzando una actividad de Bienestar"])
    ws3.append(["  con una variable de caracterización en la ficha SIRBE."])
    ws3.append([""])
    ws3.append(["  Para agregar una nueva regla: agregar una fila con el patrón a buscar"])
    ws3.append(["  en el tema/producto y la explicación que se mostrará al lector."])
    ws3.append([""])
    ws3.append(["Después de editar este Excel, correr:"])
    ws3.append(["  python scripts/generar_eje_bienestar.py"])
    ws3.column_dimensions["A"].width = 90
    for r in range(1, ws3.max_row + 1):
        ws3.cell(r, 1).font = Font(size=11, name="Calibri")
    ws3.cell(1, 1).font = Font(size=13, name="Calibri", bold=True, color="253C5C")

    RUTA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(RUTA)
    print(f"Generado: {RUTA}")
    print(f"  Exclusiones: {len(EXCLUSIONES)}")
    print(f"  Reglas de cruce: {len(REGLAS_CRUCE)}")


if __name__ == "__main__":
    main()
