# -*- coding: utf-8 -*-
"""
Actualiza la sección "Ofertas 2026" dentro de ejes/Bienestar.html a partir
del Excel de ofertas.

El script es quirúrgico: NO regenera el HTML completo. Solo reemplaza el
contenido interno del `<div class="content-section" id="ofertas_2026">`
por chips de filtros (Tipo + Clasificación) + lista de tarjetas
generadas desde el Excel. El resto del HTML (Resumen del eje, Tipos y
formatos, Datos SIRBE, header, sidebar, CSS, JS) queda intacto.

También inyecta una vez:
  - El CSS adicional de los chips y de `.oferta.oculta` dentro de <style>
  - El JS de filtrado al final del bloque <script>

Si el script se corre dos veces, detecta los marcadores y reemplaza sin
duplicar.

Fuente: ejes/oferta 2026/Oferta 2026.xlsx, hoja 'Bienestar '.
"""

import html
import re
from pathlib import Path

import openpyxl


# Marcadores adicionales para la nueva pestaña "Reporte a políticas"
MARKER_POL_INI = "<!-- ====== INICIO BLOQUE REPORTE POLITICAS GENERADO POR SCRIPT ====== -->"
MARKER_POL_FIN = "<!-- ====== FIN BLOQUE REPORTE POLITICAS GENERADO POR SCRIPT ====== -->"


_AQUI = Path(__file__).resolve().parent
BASE = _AQUI.parent
HTML_PATH = BASE / "ejes" / "Bienestar.html"
EXCEL_OFERTAS = BASE / "ejes" / "oferta 2026" / "Oferta 2026.xlsx"
# Mapeo oferta → código SIRBE (aproximación inicial, revisable a mano por
# Carolina o Diego Huertas). Columnas esperadas:
#   Nombre oferta | Tipo | Clasificación (Excel) | Código SIRBE |
#   Nombre actividad SIRBE | Curso(s) sugerido(s) en SIRBE | Notas
EXCEL_MAPEO_SIRBE = BASE / "ejes" / "oferta 2026" / "Mapeo_SIRBE_bienestar.xlsx"
# Mapeo oferta → productos de política pública (lo que Felipe pidió reportar).
# Una fila por par (oferta, producto). Se lee del Excel y se muestra debajo
# del bloque SIRBE en cada tarjeta.
EXCEL_MAPEO_POLITICAS = BASE / "ejes" / "Políticas" / "Mapeo_Politicas_bienestar.xlsx"
# Excel oficial de reportes externos (Felipe). Lo leemos para construir
# el mapeo código SIRBE → productos de política a los que aplica.
EXCEL_REPORTES_POLITICAS = BASE / "ejes" / "Políticas" / "Reportes Externos Subdirección Juventud 2026.xlsx"
# Configuración editable de exclusiones + reglas de cruce. Si el archivo
# existe, se lee de ahí; si no, se usan los valores hardcoded más abajo
# como fallback.
EXCEL_CONFIG_POLITICAS = BASE / "ejes" / "Políticas" / "Configuracion_politicas_bienestar.xlsx"

# Códigos SIRBE de Bienestar (orden de presentación: ascendente)
CODIGOS_BIENESTAR = ["511", "1485", "1486", "1487"]
NOMBRES_CODIGOS = {
    "511": "Acompañamiento y orientación psicosocial",
    "1485": "Centros de escucha",
    "1486": "Talleres informativos en prevención",
    "1487": "Cuidado frente al consumo responsable de SPA",
}

# Productos del Excel oficial que NO deben aparecer en la sección
# "Reporte a políticas" de Bienestar, aunque el Excel los liste para
# Casas de Juventud. Razón: corresponden a otros ejes (Inclusión social
# y productiva, Cultura, Liderazgo) y no al eje Bienestar.
# Se usan substrings normalizados (sin tildes, en minúsculas) sobre
# (tema, producto). Si AMBOS sub-strings aparecen, el producto se omite.
PRODUCTOS_POLITICAS_EXCLUIR = [
    ("comite intersectorial de salud", "orientacion socio-ocupacional"),
    ("sidfac", ""),
    ("formacion artistica y cultural", ""),
]


# Marcadores HTML para identificar el bloque generado y poder regenerarlo
# sin duplicar (idempotencia).
MARKER_INI = "<!-- ====== INICIO BLOQUE OFERTAS GENERADO POR SCRIPT ====== -->"
MARKER_FIN = "<!-- ====== FIN BLOQUE OFERTAS GENERADO POR SCRIPT ====== -->"

CSS_MARKER_INI = "/* === CSS chips de filtros (generado) === */"
CSS_MARKER_FIN = "/* === fin CSS chips === */"

JS_MARKER_INI = "/* === JS filtros chip (generado) === */"
JS_MARKER_FIN = "/* === fin JS filtros === */"


# ---------------------------------------------------------------------------
# Lectura del Excel
# ---------------------------------------------------------------------------

def _limpiar(valor):
    if valor is None:
        return ""
    return " ".join(str(valor).strip().split())


def _leer_mapeo_sirbe():
    """Lee Mapeo_SIRBE_bienestar.xlsx y devuelve un dict
    {nombre_oferta_normalizado: [lista_de_dicts_de_codigos]}.

    Cada oferta puede tener uno o más códigos SIRBE (una fila por par).
    Estructura del Excel:
        Nombre oferta | Código SIRBE | Nombre actividad SIRBE |
        Se usaría para | Notas
    """
    if not EXCEL_MAPEO_SIRBE.exists():
        return {}
    wb = openpyxl.load_workbook(EXCEL_MAPEO_SIRBE, data_only=True)
    ws = wb.active
    mapeo = {}
    for r in range(2, ws.max_row + 1):
        nombre = _limpiar(ws.cell(r, 1).value)
        if not nombre:
            continue
        clave = _normalizar(nombre)
        entrada = {
            "codigo": _limpiar(ws.cell(r, 2).value),
            "actividad_sirbe": _limpiar(ws.cell(r, 3).value),
            "uso": _limpiar(ws.cell(r, 4).value),
            "notas": _limpiar(ws.cell(r, 5).value),
        }
        mapeo.setdefault(clave, []).append(entrada)
    return mapeo


def _normalizar(texto):
    """Normaliza un nombre para comparar: lower, sin tildes, sin signos."""
    if not texto:
        return ""
    s = texto.lower().strip()
    for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"),
                 ("ñ", "n"), (",", ""), (".", ""), ("'", ""), ('"', ""),
                 (":", ""), (";", "")]:
        s = s.replace(a, b)
    return " ".join(s.split())


def _cargar_config_politicas():
    """Lee el Excel de configuración con exclusiones y reglas de cruce.
    Si el archivo no existe, devuelve los valores hardcoded como fallback
    para no romper la generación del HTML.

    Devuelve una tupla (exclusiones, reglas_cruce) con la misma forma que
    las constantes hardcoded de más arriba.
    """
    if not EXCEL_CONFIG_POLITICAS.exists():
        return PRODUCTOS_POLITICAS_EXCLUIR, None  # None → usa reglas hardcoded
    wb = openpyxl.load_workbook(EXCEL_CONFIG_POLITICAS, data_only=True)

    # Hoja Exclusiones: cols Tema contiene | Producto contiene | Notas
    exclusiones = []
    if "Exclusiones" in wb.sheetnames:
        ws = wb["Exclusiones"]
        for r in range(2, ws.max_row + 1):
            tema = _normalizar(_limpiar(ws.cell(r, 1).value))
            prod = _normalizar(_limpiar(ws.cell(r, 2).value))
            if tema:
                exclusiones.append((tema, prod))
    if not exclusiones:
        exclusiones = PRODUCTOS_POLITICAS_EXCLUIR

    # Hoja Reglas_cruce: cols Patrón en tema o producto | Cómo se reporta
    reglas = []
    if "Reglas_cruce" in wb.sheetnames:
        ws = wb["Reglas_cruce"]
        for r in range(2, ws.max_row + 1):
            patron = _normalizar(_limpiar(ws.cell(r, 1).value))
            explicacion = _limpiar(ws.cell(r, 2).value)
            if patron and explicacion:
                reglas.append((patron, explicacion))
    return exclusiones, (reglas if reglas else None)


def _leer_politicas_por_codigo():
    """Lee el Excel oficial de Reportes Externos y devuelve un dict
    {codigo_sirbe: [lista de productos]} donde cada código de Bienestar
    (511, 1485, 1486, 1487) lista los productos a los que aplica.

    Lee DOS hojas y fusiona/deduplica los productos:
      - 'Casas de Juventud' (la que el equipo mantiene oficialmente)
      - 'Mapeo general' filtrada por equipo responsable = "Casas de
        Juventud" (porque a veces hay productos ahí que no se copiaron
        a la hoja específica; ej.: 1.2.7 de Mujer y Equidad).

    La deduplicación es por (tipo, tema, código_producto) tomando el
    primer dígito reconocible del producto como clave de comparación.

    Detección de código SIRBE: se busca el número como palabra completa
    en la columna 'Códigos SIRBE ACTIVIDAD' (col 8).
    """
    if not EXCEL_REPORTES_POLITICAS.exists():
        return {c: [] for c in CODIGOS_BIENESTAR}
    wb = openpyxl.load_workbook(EXCEL_REPORTES_POLITICAS, data_only=True)

    # Hojas a leer; en cada una se filtra por equipo si aplica.
    hojas_y_filtro = [
        ("Casas de Juventud", None),  # toda la hoja
        ("Mapeo general", "casas de juventud"),  # solo filas del equipo
    ]

    por_codigo = {c: [] for c in CODIGOS_BIENESTAR}
    vistos_globales = {c: set() for c in CODIGOS_BIENESTAR}
    sin_codigo = []
    vistos_sin_codigo = set()

    def _clave_producto(producto):
        """Extrae los primeros caracteres significativos como clave para
        deduplicar (ej.: '1.2.7 Sensibilización...' → '1.2.7')."""
        m = re.match(r"\s*([\dA-Za-z._]+)", producto or "")
        return (m.group(1) if m else producto[:30]).strip().lower()

    # Cargar exclusiones del Excel de configuración (fallback a hardcoded)
    exclusiones, _reglas_externas = _cargar_config_politicas()

    def _excluir_producto(tema, producto):
        """Decide si un producto debe excluirse de la sección Bienestar
        porque corresponde a otro eje (Cultura, Inclusión, etc.)."""
        tema_n = _normalizar(tema)
        prod_n = _normalizar(producto)
        for sub_tema, sub_prod in exclusiones:
            if sub_tema in tema_n and (not sub_prod or sub_prod in prod_n):
                return True
            if sub_tema in prod_n and (not sub_prod or sub_prod in prod_n):
                return True
        return False

    for nombre_hoja, filtro_equipo in hojas_y_filtro:
        if nombre_hoja not in wb.sheetnames:
            continue
        ws = wb[nombre_hoja]
        for r in range(2, ws.max_row + 1):
            tipo = _limpiar(ws.cell(r, 1).value)
            tema = _limpiar(ws.cell(r, 2).value)
            producto = _limpiar(ws.cell(r, 3).value)
            equipo = _limpiar(ws.cell(r, 4).value).lower()
            cod_actividad = str(ws.cell(r, 8).value or "")
            if not producto:
                continue
            if filtro_equipo and filtro_equipo not in equipo:
                continue
            # Excluir productos que corresponden a otros ejes
            if _excluir_producto(tema, producto):
                continue
            # ¿Tiene algún código de Bienestar?
            tiene_alguno = False
            for c in CODIGOS_BIENESTAR:
                if re.search(r"\b" + c + r"\b", cod_actividad):
                    tiene_alguno = True
                    clave = (tema.lower(), _clave_producto(producto))
                    if clave in vistos_globales[c]:
                        continue
                    vistos_globales[c].add(clave)
                    por_codigo[c].append({
                        "tipo": tipo,
                        "tema": tema,
                        "producto": producto,
                    })
            # Si no tiene código de Bienestar, va a la lista de "sin código"
            if not tiene_alguno:
                clave = (tema.lower(), _clave_producto(producto))
                if clave not in vistos_sin_codigo:
                    vistos_sin_codigo.add(clave)
                    sin_codigo.append({
                        "tipo": tipo,
                        "tema": tema,
                        "producto": producto,
                    })
    return por_codigo, sin_codigo


def _html_seccion_politicas():
    """Genera el HTML del contenido de la pestaña 'Reporte a políticas':
    por cada código SIRBE de Bienestar, lista los productos que lo
    incluyen en su columna 'Códigos SIRBE ACTIVIDAD'. Al final agrega
    una sección con los productos que no tienen código asignado (col '-')
    para que sea visible qué falta llenar en el Excel oficial."""
    por_codigo, sin_codigo = _leer_politicas_por_codigo()

    # Descripciones cortas de cada código SIRBE para que cada tarjeta
    # sea autocontenida (no obligue al lector a leer la sección de
    # Ofertas 2026 para entender qué es cada código).
    descripciones_codigo = {
        "511": "Atenciones individuales con acompañamiento y orientación psicosocial al joven.",
        "1485": "Espacios estructurados de escucha sobre DSDR. Incluye formación, acceso a información y ejercicio de derechos. Las salas de escucha de las Casas se reportan aquí.",
        "1486": "Sesiones grupales con propósito formativo sobre DSDR, salud mental, PVBG y SPA. La mayoría de los ciclos, experiencias y encuentros caen aquí.",
        "1487": "Espacios específicamente enfocados en prevención del consumo de sustancias psicoactivas, con metodología especializada.",
    }

    bloques = []
    for codigo in CODIGOS_BIENESTAR:
        productos = por_codigo.get(codigo, [])
        nombre = NOMBRES_CODIGOS.get(codigo, "")
        descripcion = descripciones_codigo.get(codigo, "")

        # Agrupar productos por tema (política / plan / programa) para
        # listar bajo un solo título de política los varios productos que
        # se reportan bajo ese código.
        por_tema = {}
        orden_temas = []
        for p in productos:
            clave = (p["tipo"], p["tema"])
            if clave not in por_tema:
                por_tema[clave] = []
                orden_temas.append(clave)
            por_tema[clave].append(p["producto"])

        if productos:
            grupos_html = []
            for tipo, tema in orden_temas:
                lista_prods = "".join(
                    f'                                <li class="rp-prod">{_esc(prod)}</li>\n'
                    for prod in por_tema[(tipo, tema)]
                )
                grupos_html.append(
                    '                        <div class="rp-grupo">\n'
                    f'                            <p class="rp-grupo-tipo">{_esc(tipo)}</p>\n'
                    f'                            <p class="rp-grupo-tema">{_esc(tema)}</p>\n'
                    '                            <ul class="rp-prods">\n'
                    f'{lista_prods}'
                    '                            </ul>\n'
                    '                        </div>'
                )
            derecha = "\n".join(grupos_html)
        else:
            derecha = (
                '                        <p class="rp-vacio">Ningún producto del Excel oficial '
                'menciona este código en su columna "Códigos SIRBE ACTIVIDAD".</p>'
            )

        bloques.append(
            '                <article class="rp-card">\n'
            '                    <div class="rp-card-codigo">\n'
            '                        <p class="rp-card-label">Código SIRBE</p>\n'
            f'                        <p class="rp-card-num">{codigo}</p>\n'
            f'                        <p class="rp-card-nombre">{_esc(nombre)}</p>\n'
            f'                        <p class="rp-card-desc">{_esc(descripcion)}</p>\n'
            '                    </div>\n'
            '                    <div class="rp-card-grupos">\n'
            f'{derecha}\n'
            '                    </div>\n'
            '                </article>'
        )

    intro = (
        '                <p style="font-family:Figtree, sans-serif; font-size:0.92rem; color:#3a3a3a; line-height:1.65; margin: 8px 0 14px; max-width:820px;">'
        'El eje de Bienestar del servicio Casas de Juventud apunta de manera directa y '
        'transversal a varias Políticas Públicas Distritales y Planes Intersectoriales, '
        'gracias a sus acciones en Derechos Sexuales y Reproductivos (DSYR), Salud Mental, '
        'Prevención de Violencias Basadas en Género y prevención de Sustancias Psicoactivas (SPA).'
        '</p>\n'
        '                <p style="font-family:Figtree, sans-serif; font-size:0.92rem; color:#3a3a3a; line-height:1.65; margin: 0 0 14px; max-width:820px;">'
        'Al ser un servicio de carácter inclusivo, todas las ofertas preventivas y de '
        'acompañamiento en bienestar apuntan de forma transversal a las metas de '
        'vinculación con enfoque diferencial dentro de las Políticas Públicas de '
        'Ruralidad, Discapacidad, Migrantes Internacionales, Pueblos Indígenas, '
        'Pueblo Raizal, Rrom, y comunidades Negras, Afrocolombianas y Palenqueras.'
        '</p>\n'
        '                <p style="font-family:Figtree, sans-serif; font-size:0.92rem; color:#3a3a3a; line-height:1.65; margin: 0 0 14px; max-width:820px;">'
        'Cada atención que entrega Casas de Juventud se registra en <strong>SIRBE</strong> '
        '(Sistema de Registro de Beneficiarios) bajo un <strong>código de actividad</strong> '
        'que indica qué se hizo. El eje Bienestar usa cuatro códigos: 511, 1485, 1486 y 1487. '
        'Cada tarjeta de esta sección muestra un código y los productos de política a los que '
        'reporta cuando se aplica esa actividad.'
        '</p>\n'
        '                <p style="font-family:Figtree, sans-serif; font-size:0.92rem; color:#3a3a3a; line-height:1.65; margin: 0 0 24px; max-width:820px;">'
        'El sistema de reporte funciona como una <strong>matriz cruzada</strong>: el código de actividad de SIRBE indica <em>qué</em> se hizo, y las variables de caracterización del participante en la ficha SIRBE indican <em>con quién</em> se hizo. Por eso muchos productos y políticas poblacionales no necesitan un código exclusivo: se alimentan de las actividades generales de Bienestar cruzadas con la caracterización del joven.'
        '</p>'
    )

    # Sección final con los productos que no tienen código de actividad
    # SIRBE propio. NO son huérfanos: se reportan mediante el cruce de la
    # actividad ejecutada (bajo cualquiera de los 4 códigos de Bienestar)
    # con las variables de caracterización del participante en la ficha
    # FOR-PSS-328 (grupo étnico, rol ocupacional, víctima de conflicto,
    # ubicación territorial, etc.) o con marcadores en el nombre del
    # curso. Aquí se documenta cómo se reporta cada uno.
    sin_codigo_html = ""
    if sin_codigo:
        # Reglas de cruce: nombre normalizado del tema → cómo se reporta.
        # Se buscan coincidencias parciales para tolerar variaciones de
        # texto en el Excel oficial.
        # Si hay reglas en el Excel de configuración, usarlas; si no,
        # caer al fallback hardcoded de abajo.
        _exclusiones_ext, reglas_excel = _cargar_config_politicas()
        reglas_cruce_fallback = [
            ("lgbti", "Ficha SIRBE casillas E (identidad de género), F (identidad transgénero) y G (orientación sexual). El equipo de la DADE filtra a los y las jóvenes que se autorreconocen en sectores LGBTI."),
            ("negra", "Ficha SIRBE casilla AA (grupo étnico). El equipo de la DADE toma a los y las jóvenes que reportaron autorreconocimiento étnico."),
            ("afrocolomb", "Ficha SIRBE casilla AA (grupo étnico). El equipo de la DADE toma a los y las jóvenes que reportaron autorreconocimiento étnico."),
            ("indígena", "Ficha SIRBE casillas AA (grupo étnico) y AB (grupo indígena específico)."),
            ("indigena", "Ficha SIRBE casillas AA (grupo étnico) y AB (grupo indígena específico)."),
            ("raizal", "Ficha SIRBE casilla AA (grupo étnico, opción Raizal)."),
            ("rrom", "Ficha SIRBE casilla AA (grupo étnico, opción Rrom / Gitano)."),
            ("actividades sexuales pagadas", "Ficha SIRBE casilla T (Rol Ocupacional, opción Actividades Sexuales pagadas)."),
            ("víctimas", "Cruce interno de la DADE con el Registro Único de Víctimas (RUV) y la ficha SIRBE casilla AC (víctima de conflicto armado)."),
            ("victimas", "Cruce interno de la DADE con el Registro Único de Víctimas (RUV) y la ficha SIRBE casilla AC (víctima de conflicto armado)."),
            ("pdet", "Filtro por ubicación territorial: actividades prestadas en la Localidad 20 de Sumapaz o en zonas rurales priorizadas PDET."),
            ("reincorporación", "Al diligenciar el nombre del curso en SIRBE se agrega el marcador “Reincorporados”, y el equipo filtra por ese marcador."),
            ("reincorporacion", "Al diligenciar el nombre del curso en SIRBE se agrega el marcador “Reincorporados”, y el equipo filtra por ese marcador."),
            ("ruralidad", "Ficha SIRBE casilla AH (tipo de predio: rural)."),
            ("discapacidad", "Ficha SIRBE casilla Z (Persona con discapacidad) y tipo de discapacidad."),
            ("migrantes", "Ficha SIRBE casilla A (tipo de documento), opciones de cédula de extranjería, permisos o pasaporte."),
        ]

        # Usar reglas del Excel si las hay; si no, fallback hardcoded
        reglas_cruce = reglas_excel if reglas_excel else reglas_cruce_fallback

        def _explicacion_cruce(producto, tema):
            texto = _normalizar(producto + " " + tema)
            for clave, exp in reglas_cruce:
                if clave in texto:
                    return exp
            return None

        # Agrupar productos sin código también por tipo + tema
        por_tema_sc = {}
        orden_temas_sc = []
        for p in sin_codigo:
            clave = (p["tipo"], p["tema"])
            if clave not in por_tema_sc:
                por_tema_sc[clave] = []
                orden_temas_sc.append(clave)
            por_tema_sc[clave].append(p["producto"])

        grupos_sc = []
        for tipo, tema in orden_temas_sc:
            exp = _explicacion_cruce(" ".join(por_tema_sc[(tipo, tema)]), tema)
            cruce_html = ""
            if exp:
                cruce_html = f'\n                            <p class="rp-cruce"><strong>Cómo se reporta:</strong> {exp}</p>'
            lista_prods = "".join(
                f'                                <li class="rp-prod">{_esc(prod)}</li>\n'
                for prod in por_tema_sc[(tipo, tema)]
            )
            grupos_sc.append(
                '                        <div class="rp-grupo">\n'
                f'                            <p class="rp-grupo-tipo">{_esc(tipo)}</p>\n'
                f'                            <p class="rp-grupo-tema">{_esc(tema)}</p>\n'
                '                            <ul class="rp-prods">\n'
                f'{lista_prods}'
                '                            </ul>'
                f'{cruce_html}\n'
                '                        </div>'
            )
        grupos_html_sc = "\n".join(grupos_sc)

        sin_codigo_html = (
            '\n\n                <article class="rp-card rp-card-cruce">\n'
            '                    <div class="rp-card-codigo">\n'
            '                        <p class="rp-card-label">Sin código SIRBE propio</p>\n'
            '                        <p class="rp-card-num">+</p>\n'
            '                        <p class="rp-card-nombre">Productos que se reportan por cruce de variables</p>\n'
            '                        <p class="rp-aclaracion">Estos productos no tienen un código de actividad SIRBE propio. '
            'El reporte se logra cruzando las atenciones registradas bajo cualquiera de los cuatro códigos de Bienestar '
            'con las variables de caracterización del participante en la ficha SIRBE (grupo étnico, orientación sexual, '
            'rol ocupacional, ubicación territorial, condición de víctima, etc.) o con marcadores específicos en el '
            'nombre del curso.</p>\n'
            '                    </div>\n'
            '                    <div class="rp-card-grupos">\n'
            f'{grupos_html_sc}\n'
            '                    </div>\n'
            '                </article>'
        )

    return (
        MARKER_POL_INI + "\n" + intro + "\n\n"
        + "\n\n".join(bloques)
        + sin_codigo_html
        + "\n                " + MARKER_POL_FIN
    )


def _leer_ofertas():
    if not EXCEL_OFERTAS.exists():
        return []
    wb = openpyxl.load_workbook(EXCEL_OFERTAS, data_only=True)
    nombre_hoja = next(
        (n for n in wb.sheetnames if "ienestar" in n.lower() and "alianza" not in n.lower()),
        None,
    )
    if nombre_hoja is None:
        return []
    ws = wb[nombre_hoja]

    merge_map = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                merge_map[(r, c)] = top_val

    def cell(r, c):
        v = ws.cell(r, c).value
        if v is None:
            v = merge_map.get((r, c))
        return v

    por_clave, orden = {}, []
    for r in range(3, ws.max_row + 1):
        nombre = _limpiar(cell(r, 5))
        if not nombre:
            continue
        # Exclusiones editoriales
        if _normalizar(nombre) in OFERTAS_EXCLUIR:
            continue
        tipo = _limpiar(cell(r, 3))
        clasif = _limpiar(cell(r, 4)).split("\n")[0].strip()
        # Aplicar override editorial si existe para esta oferta
        override = OFERTAS_OVERRIDE.get(_normalizar(nombre), {})
        if "tipo" in override:
            tipo = override["tipo"]
        if "clasificacion" in override:
            clasif = override["clasificacion"]

        clave = (tipo, clasif, nombre)
        if clave not in por_clave:
            descripcion = override.get("descripcion") or _limpiar(cell(r, 6))
            por_clave[clave] = {
                "tipo": tipo,
                "clasificacion": clasif,
                "nombre": nombre,
                "modalidad": _limpiar(cell(r, 2)),
                "descripcion": descripcion,
                "implementa": _limpiar(cell(r, 7)),
                "dirigido_a": _limpiar(cell(r, 8)),
                "localidad": _limpiar(cell(r, 9)),
                "mes": _limpiar(cell(r, 10)),
                "meta": _limpiar(cell(r, 11)),
            }
            orden.append(clave)
    return [por_clave[k] for k in orden]


# ---------------------------------------------------------------------------
# Helpers HTML
# ---------------------------------------------------------------------------

def _slug(texto):
    if not texto:
        return "sin-clasif"
    s = texto.lower().strip()
    for a, b in [
        ("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"),
        ("ñ", "n"), (" ", "-"), ("/", "-"), (",", ""), ("(", ""),
        (")", ""), ("'", ""), ('"', ""), (":", ""), (".", ""),
    ]:
        s = s.replace(a, b)
    while "--" in s:
        s = s.replace("--", "-")
    return s.strip("-") or "sin-clasif"


def _esc(texto):
    return html.escape(texto or "")


# Color de título por tipo de oferta (estilo banderas usado en Tipos y formatos).
COLOR_TIPO = {
    "Fortalecimiento de habilidades": "#f4676e",
    "Conexión con oportunidades": "#1eaf76",
    "Participación": "#2fa4d4",
}

# Ofertas que NO deben aparecer en la página (decisión editorial de Carolina).
# Se comparan por nombre normalizado, así que cualquier variación tipográfica
# (mayúsculas, tildes) se ignora.
OFERTAS_EXCLUIR = {
    "fondos de desarrollo local",
}

# Overrides editoriales por oferta: campos que reemplazan o complementan
# lo que viene del Excel. La clave es el nombre normalizado. Útil cuando
# el Excel trae datos genéricos o cuando una oferta no encaja en los tipos
# estándar (ej. Semana Andina, que no es Fortalecimiento / Conexión /
# Participación: es un evento conmemorativo aparte).
OFERTAS_OVERRIDE = {
    "semana andina": {
        "tipo": "Evento",
        "clasificacion": "",
        "descripcion": (
            "Conmemoración anual liderada por el Ministerio de Salud y entidades "
            "distritales y nacionales (también articulada en Bolivia, Chile, "
            "Ecuador, Perú y Venezuela). Busca la prevención del embarazo en la "
            "infancia y la adolescencia y la promoción de los derechos sexuales y "
            "reproductivos de las y los jóvenes. Bogotá la articula con las Casas "
            "de Juventud mediante concursos artísticos con incentivos económicos "
            "(freestyle, slam poético, ilustración, cortometrajes, sexpo-ferias), "
            "tomas de colegios y SENA, foros académicos y festivales de cierre. "
            "Reúne ofertas de varios formatos del eje. Lemas recientes: "
            "‘Decidir mi futuro es mi derecho’ (2025), ‘Oportunidades que "
            "transforman’ (2024), ‘En las familias hablamos claro’ (2023)."
        ),
    },
}


def _tarjeta(o, mapeo_sirbe):
    """Genera una tarjeta horizontal de oferta. Mismo aire que las fichas
    'A tener en cuenta' de Forjar pero ancha en lugar de en grid de 3.
    Layout:
        [ Clasif + Título + Descripción ] [ Bloque SIRBE a la derecha ]
    El bloque SIRBE muestra el código y el nombre de la actividad bajo la
    cual se reporta esta oferta en SIRBE (aproximación del mapeo)."""
    tipo_slug = _slug(o["tipo"]) if o["tipo"] else "sin-tipo"
    clasif_slug = _slug(o["clasificacion"]) if o["clasificacion"] else "sin-clasif"
    color = COLOR_TIPO.get(o["tipo"], "#2F3E3C")
    descripcion = _esc(o["descripcion"]) or "<em style=\"color:#888;\">Sin descripción registrada en el Excel.</em>"
    clasif_label = _esc(o["clasificacion"])
    tipo_label = _esc(o["tipo"])
    # Tipo + clasificación van combinados en una sola etiqueta gris
    # encima del nombre (kicker). Solo el nombre lleva el color del tipo.
    etiqueta_partes = [x for x in (tipo_label, clasif_label) if x]
    etiqueta_html = ""
    if etiqueta_partes:
        etiqueta_html = (
            f'<p class="bn-kicker">{" &middot; ".join(etiqueta_partes)}</p>'
        )

    # Buscar mapeo SIRBE para esta oferta. Primero match exacto;
    # si no, buscar por contención (Excel puede tener texto extra
    # tipo "(No es prioritaria)" que el mapeo no incluye).
    clave = _normalizar(o["nombre"])
    entradas = mapeo_sirbe.get(clave, [])
    if not entradas:
        for k, v in mapeo_sirbe.items():
            if k and (k in clave or clave in k):
                entradas = v
                break

    entradas_validas = [e for e in entradas if e.get("codigo") or e.get("actividad_sirbe")]

    if entradas_validas:
        items = []
        for e in entradas_validas:
            uso_html = ""
            if e.get("uso"):
                uso_html = f'<p class="bn-sirbe-uso">{_esc(e["uso"])}</p>'
            items.append(
                '                            <div class="bn-sirbe-item">\n'
                f'                                <p class="bn-sirbe-codigo">{_esc(e["codigo"]) or "—"}</p>\n'
                f'                                <p class="bn-sirbe-actividad">{_esc(e["actividad_sirbe"])}</p>\n'
                f'                                {uso_html}\n'
                '                            </div>'
            )
        items_html = "\n".join(items)
        sirbe_html = f'''
                        <aside class="bn-sirbe">
                            <p class="bn-sirbe-label">Reporte SIRBE</p>
{items_html}
                        </aside>'''
    else:
        sirbe_html = '''
                        <aside class="bn-sirbe">
                            <p class="bn-sirbe-label">Reporte SIRBE</p>
                            <p class="bn-sirbe-pendiente">Por mapear</p>
                        </aside>'''

    # Mini-ficha de metadatos: solo los campos con valor, en lista
    # label/valor estilo editorial (sin badges, sin cápsulas).
    meta_items = []
    for label, key in [
        ("Modalidad", "modalidad"),
        ("Dirigido a", "dirigido_a"),
        ("Mes", "mes"),
        ("Meta", "meta"),
    ]:
        v = o.get(key, "")
        if v:
            meta_items.append(f'                                <dt>{label}</dt><dd>{_esc(v)}</dd>')
    meta_html = ""
    if meta_items:
        meta_html = (
            '\n                            <dl class="bn-meta">\n'
            + "\n".join(meta_items) + "\n"
            '                            </dl>'
        )

    return f'''                    <div class="bn-card" data-tipo="{tipo_slug}" data-clasif="{clasif_slug}">
                        <div class="bn-card-main">
                            {etiqueta_html}
                            <h4 class="bn-titulo" style="color:{color};">{_esc(o["nombre"])}</h4>
                            <p class="bn-texto">{descripcion}</p>{meta_html}
                        </div>{sirbe_html}
                    </div>'''


INTRO_OFERTAS_SIRBE = """\
                <div class="bn-intro">
                    <p>Las ofertas del eje Bienestar son <strong>actividades puntuales y procesos pedag&oacute;gicos</strong> que se entregan en las Casas de Juventud. Cada una se registra en <strong>SIRBE</strong> (Sistema de Registro de Beneficiarios) bajo uno de los cuatro c&oacute;digos de actividad que la subdirecci&oacute;n usa para las atenciones de Casas de Juventud del eje Bienestar:</p>

                    <div class="bn-codigos-lista">
                        <div class="bn-codigo-item">
                            <p class="bn-codigo-num">511</p>
                            <div class="bn-codigo-cuerpo">
                                <p class="bn-codigo-nombre">Acompa&ntilde;amiento y orientaci&oacute;n psicosocial</p>
                                <p class="bn-codigo-desc">Atenciones individuales con acompa&ntilde;amiento y orientaci&oacute;n psicosocial al joven.</p>
                            </div>
                        </div>
                        <div class="bn-codigo-item">
                            <p class="bn-codigo-num">1485</p>
                            <div class="bn-codigo-cuerpo">
                                <p class="bn-codigo-nombre">Centros de escucha</p>
                                <p class="bn-codigo-desc">Espacios estructurados de escucha sobre derechos sexuales y reproductivos. Incluye formaci&oacute;n, acceso a informaci&oacute;n y ejercicio de derechos. Aqu&iacute; se encuentran las salas de escucha de las Casas.</p>
                            </div>
                        </div>
                        <div class="bn-codigo-item">
                            <p class="bn-codigo-num">1486</p>
                            <div class="bn-codigo-cuerpo">
                                <p class="bn-codigo-nombre">Talleres informativos en prevenci&oacute;n</p>
                                <p class="bn-codigo-desc">Sesiones grupales con prop&oacute;sito formativo sobre DSDR, salud mental, PVBG y SPA. Es el c&oacute;digo donde caen la mayor&iacute;a de los ciclos, experiencias, encuentros y eventos masivos.</p>
                            </div>
                        </div>
                        <div class="bn-codigo-item">
                            <p class="bn-codigo-num">1487</p>
                            <div class="bn-codigo-cuerpo">
                                <p class="bn-codigo-nombre">Cuidado frente al consumo responsable de SPA</p>
                                <p class="bn-codigo-desc">Espacios espec&iacute;ficamente enfocados en prevenci&oacute;n del consumo de sustancias psicoactivas, con metodolog&iacute;a especializada y enfoque de reducci&oacute;n de da&ntilde;o.</p>
                            </div>
                        </div>
                    </div>

                    <p class="bn-intro-nota">El c&oacute;digo <strong>1599 Orientaci&oacute;n psicosocial</strong> estaba antes en la tabla, pero ya no aparece en la tabla de homologaci&oacute;n vigente.</p>

                    <p class="bn-intro-nota">Para mayor informaci&oacute;n sobre la desagregaci&oacute;n SIRBE para el eje Casas de Juventud &mdash; Bienestar, consulte la <a href="#datos_sirbe" onclick="showContent('datos_sirbe'); return false;">desagregaci&oacute;n detallada de Datos SIRBE 2025</a>.</p>
                </div>

"""


def _bloque_ofertas(ofertas, mapeo_sirbe):
    """Genera el contenido completo a inyectar entre los marcadores.
    Incluye un intro explicando los 5 códigos SIRBE de Bienestar más la
    lista de tarjetas. Los filtros se removieron por decisión de Carolina."""
    tarjetas = "\n\n".join(_tarjeta(o, mapeo_sirbe) for o in ofertas)
    return f'''{MARKER_INI}
{INTRO_OFERTAS_SIRBE}
                <div class="bn-grid">
{tarjetas}
                </div>
                {MARKER_FIN}'''


# ---------------------------------------------------------------------------
# CSS y JS que se inyectan una vez en el HTML
# ---------------------------------------------------------------------------

CSS_CHIPS = f"""
{CSS_MARKER_INI}
/* Intro de la sección Ofertas 2026: párrafo + lista de 5 códigos SIRBE
   con esbozo + nota a Datos SIRBE 2025. Sin bandas, sin sombras
   dramáticas; jerarquía dada por la tipografía y el espaciado. */
.bn-intro {{ margin-bottom: 32px; }}
.bn-intro > p {{ font-family: 'Figtree', sans-serif; font-size: 0.92rem; color: #3a3a3a; line-height: 1.65; margin: 0 0 18px; max-width: 820px; }}
.bn-codigos-lista {{ display: flex; flex-direction: column; gap: 14px; margin: 20px 0 22px; }}
.bn-codigo-item {{ display: grid; grid-template-columns: 80px 1fr; gap: 18px; align-items: start; padding: 14px 0; }}
.bn-codigo-item + .bn-codigo-item {{ padding-top: 14px; }}
.bn-codigo-num {{ font-family: 'Anton', 'Segoe UI', sans-serif; font-size: 2.1rem; color: #253C5C; margin: 0; line-height: 1; letter-spacing: 0.01em; }}
.bn-codigo-cuerpo {{ min-width: 0; }}
.bn-codigo-nombre {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; font-size: 0.86rem; color: #2F3E3C; letter-spacing: 0.03em; margin: 0 0 5px; line-height: 1.3; }}
.bn-codigo-desc {{ font-family: 'Figtree', sans-serif; font-size: 0.86rem; color: #555; line-height: 1.55; margin: 0; }}
.bn-codigo-desc strong {{ color: #2F3E3C; font-weight: 700; }}
.bn-intro-nota {{ font-family: 'Figtree', sans-serif; font-size: 0.86rem; color: #555; line-height: 1.6; margin: 18px 0 0; padding-top: 14px; max-width: 820px; }}
.bn-intro-nota a {{ color: #253C5C; font-weight: 600; text-decoration: underline; text-underline-offset: 3px; }}
@media (max-width: 720px) {{
    .bn-codigo-item {{ grid-template-columns: 1fr; gap: 4px; }}
    .bn-codigo-num {{ font-size: 1.7rem; }}
}}

/* Tarjetas horizontales de ofertas Bienestar: contenido principal a la
   izquierda, bloque SIRBE con el código de reporte a la derecha. Fondo
   crema, sin bandas. Hover sutil. */
.bn-grid {{ display: flex; flex-direction: column; gap: 14px; margin-top: 8px; }}
.bn-card {{ background: #f5efd2; border-radius: 12px; padding: 20px 22px; display: grid; grid-template-columns: 1fr 280px; gap: 24px; align-items: stretch; box-shadow: 0 2px 8px rgba(0,0,0,0.04); transition: transform 0.18s ease, box-shadow 0.18s ease; }}
.bn-card:hover {{ transform: translateY(-2px); box-shadow: 0 6px 18px rgba(0,0,0,0.07); }}
.bn-card.oculta {{ display: none; }}
.bn-card-main {{ min-width: 0; }}
.bn-kicker {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; letter-spacing: 0.07em; font-size: 0.7rem; color: #888; margin: 0 0 8px; line-height: 1.3; }}
.bn-titulo {{ font-family: 'Antonio', 'Anton', 'Segoe UI', sans-serif; font-weight: 700; font-size: 1.05rem; text-transform: uppercase; letter-spacing: 0.02em; margin: 0 0 10px; line-height: 1.22; }}
.bn-texto {{ font-family: 'Figtree', 'Segoe UI', sans-serif; font-weight: 500; font-size: 0.83rem; color: #3a3a3a; line-height: 1.6; margin: 0; }}
.bn-texto strong {{ font-weight: 700; color: #2f3e3c; }}

/* Mini-ficha de metadatos al pie de la tarjeta: lista label/valor en
   dos columnas, estilo editorial. Etiqueta en Antonio Bold pequeña,
   valor en Figtree. Sin cápsulas, sin colores de fondo. */
.bn-meta {{ display: grid; grid-template-columns: 110px 1fr; gap: 4px 14px; margin: 14px 0 0; padding-top: 12px; font-size: 0.8rem; }}
.bn-meta dt {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; color: #2F3E3C; text-transform: uppercase; letter-spacing: 0.05em; font-size: 0.7rem; padding-top: 2px; }}
.bn-meta dd {{ color: #555; margin: 0; line-height: 1.5; }}
@media (max-width: 600px) {{
    .bn-meta {{ grid-template-columns: 1fr; gap: 1px; }}
    .bn-meta dt {{ padding-top: 8px; }}
    .bn-meta dt:first-of-type {{ padding-top: 2px; }}
}}

/* Bloque SIRBE a la derecha de cada tarjeta: sin caja, sin bordes ni
   bandas; la separación visual la da el grid y la tipografía. */
.bn-sirbe {{ padding: 4px 0; display: flex; flex-direction: column; justify-content: center; gap: 14px; }}
.bn-sirbe-label {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; font-size: 0.66rem; color: #888; margin: 0; }}
.bn-sirbe-item {{ margin: 0; }}
.bn-sirbe-codigo {{ font-family: 'Anton', 'Segoe UI', sans-serif; font-size: 1.55rem; color: #253C5C; margin: 0 0 4px; line-height: 1; letter-spacing: 0.01em; }}
.bn-sirbe-actividad {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; font-size: 0.72rem; color: #2F3E3C; letter-spacing: 0.03em; margin: 0 0 4px; line-height: 1.3; }}
.bn-sirbe-uso {{ font-family: 'Figtree', 'Segoe UI', sans-serif; font-size: 0.78rem; color: #555; line-height: 1.5; margin: 0; }}
.bn-sirbe-pendiente {{ font-family: 'Figtree', 'Segoe UI', sans-serif; font-size: 0.82rem; color: #aaa; margin: 0; }}

@media (max-width: 900px) {{
    .bn-card {{ grid-template-columns: 1fr; }}
    .bn-sirbe {{ padding-top: 10px; margin-top: 6px; }}
}}

/* Sección "Reporte a políticas": tarjetas horizontales tipo bn-card.
   Mismo lenguaje visual que las tarjetas de "Ofertas 2026" pero con
   datos invertidos: izquierda = código SIRBE, derecha = productos
   agrupados por política. */
.rp-card {{ background: #f5efd2; border-radius: 12px; padding: 22px 24px; display: grid; grid-template-columns: 1fr 2fr; gap: 28px; align-items: start; margin-bottom: 16px; box-shadow: 0 2px 8px rgba(0,0,0,0.04); transition: transform 0.18s ease, box-shadow 0.18s ease; }}
.rp-card:hover {{ transform: translateY(-2px); box-shadow: 0 6px 18px rgba(0,0,0,0.07); }}
.rp-card-codigo {{ min-width: 0; }}
.rp-card-label {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; font-size: 0.66rem; color: #888; margin: 0 0 6px; }}
.rp-card-num {{ font-family: 'Anton', 'Segoe UI', sans-serif; font-size: 2.4rem; color: #253C5C; margin: 0 0 6px; line-height: 1; letter-spacing: 0.01em; }}
.rp-card-nombre {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; font-size: 0.88rem; color: #2F3E3C; letter-spacing: 0.03em; margin: 0 0 10px; line-height: 1.3; }}
.rp-card-desc {{ font-family: 'Figtree', sans-serif; font-size: 0.83rem; color: #555; line-height: 1.55; margin: 0; }}
.rp-card-grupos {{ display: flex; flex-direction: column; gap: 18px; min-width: 0; }}
.rp-grupo {{ margin: 0; }}
.rp-grupo-tipo {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em; font-size: 0.66rem; color: #888; margin: 0 0 2px; }}
.rp-grupo-tema {{ font-family: 'Antonio', 'Anton', 'Figtree', sans-serif; font-weight: 700; text-transform: uppercase; font-size: 0.86rem; color: #253C5C; letter-spacing: 0.02em; margin: 0 0 8px; line-height: 1.3; }}
.rp-prods {{ list-style: none; padding: 0; margin: 0; display: flex; flex-direction: column; gap: 6px; }}
.rp-prod {{ font-family: 'Figtree', sans-serif; font-size: 0.84rem; color: #3a3a3a; line-height: 1.55; padding-left: 14px; position: relative; }}
.rp-prod::before {{ content: '·'; position: absolute; left: 0; top: -2px; color: #253C5C; font-weight: 700; font-size: 1.2rem; line-height: 1; }}
.rp-vacio {{ font-family: 'Figtree', sans-serif; font-size: 0.85rem; color: #aaa; font-style: italic; margin: 0; }}

/* Tarjeta "Productos que se reportan por cruce de variables": misma
   estética pero con cruce explicado bajo cada producto. */
.rp-card-cruce {{ grid-template-columns: 1fr; }}
.rp-card-cruce .rp-card-codigo {{ border-bottom: 1px solid rgba(47,62,60,0.12); padding-bottom: 14px; margin-bottom: 4px; }}
.rp-card-cruce .rp-card-num {{ font-size: 1.6rem; color: #555; }}
.rp-aclaracion {{ font-family: 'Figtree', sans-serif; font-size: 0.85rem; color: #555; line-height: 1.6; margin: 0 0 12px; max-width: 760px; }}
.rp-aclaracion strong, .rp-aclaracion em {{ color: #253C5C; }}
.rp-cruce {{ font-family: 'Figtree', sans-serif; font-size: 0.8rem; color: #555; line-height: 1.55; margin: 6px 0 12px 14px; padding: 8px 12px; background: rgba(255,255,255,0.55); border-radius: 6px; }}
.rp-cruce strong {{ color: #253C5C; font-weight: 700; }}
.rp-cruce em {{ color: #2F3E3C; font-style: italic; }}

@media (max-width: 720px) {{
    .rp-card {{ grid-template-columns: 1fr; gap: 14px; }}
}}
@media (max-width: 600px) {{
    .rp-codigo-header, .rp-item {{ grid-template-columns: 1fr; gap: 4px; }}
    .rp-producto {{ grid-column: 1; }}
    .rp-vacio {{ padding-left: 0; }}
}}

/* Fix overflow de la sección Datos SIRBE 2025: las tablas y el árbol de
   cursos tienen textos largos que sobrepasan el ancho del .card. Esto
   confina las tablas a su contenedor y permite scroll horizontal si hace
   falta. También fuerza word-wrap en celdas. */
#datos_sirbe .card {{ overflow-x: auto; }}
#datos_sirbe table {{ table-layout: auto; max-width: 100%; }}
#datos_sirbe table td, #datos_sirbe table th {{ word-break: break-word; overflow-wrap: break-word; }}
#datos_sirbe [id^="cursos_BIENESTAR_"] {{ overflow-x: auto; }}
{CSS_MARKER_FIN}
"""

# Bloque JS vacío: ya no usamos filtros, pero mantenemos los marcadores
# para que regeneraciones futuras puedan reemplazar este bloque sin dejar
# residuos del script anterior.
JS_FILTROS = f"\n{JS_MARKER_INI}\n{JS_MARKER_FIN}\n"

# Bloque legacy (referencia, no se usa). Lo dejo comentado para recordar
# qué hacía si alguien necesita restaurar los filtros.
_JS_FILTROS_LEGACY = f"""
{JS_MARKER_INI}
document.querySelectorAll('.chip').forEach(function(btn) {{
    btn.addEventListener('click', function(e) {{
        e.stopPropagation();
        var grupo = btn.hasAttribute('data-tipo') ? 'data-tipo' : 'data-clasif';
        var valor = btn.getAttribute(grupo);
        if (valor === 'todos') {{
            document.querySelectorAll('.chip[' + grupo + ']').forEach(function(c) {{ c.classList.remove('activo'); }});
            btn.classList.add('activo');
        }} else {{
            btn.classList.toggle('activo');
            var todos = document.querySelector('.chip[' + grupo + '="todos"]');
            var activos = document.querySelectorAll('.chip[' + grupo + ']:not([' + grupo + '="todos"]).activo');
            if (activos.length === 0) {{ todos.classList.add('activo'); }} else {{ todos.classList.remove('activo'); }}
        }}
        aplicarFiltrosOfertas();
    }});
}});

function aplicarFiltrosOfertas() {{
    var tiposActivos = Array.from(document.querySelectorAll('.chip[data-tipo].activo')).map(function(c) {{ return c.getAttribute('data-tipo'); }});
    var clasifsActivas = Array.from(document.querySelectorAll('.chip[data-clasif].activo')).map(function(c) {{ return c.getAttribute('data-clasif'); }});
    var tipoTodos = tiposActivos.indexOf('todos') !== -1;
    var clasifTodas = clasifsActivas.indexOf('todos') !== -1;
    var visibles = 0;
    document.querySelectorAll('#ofertas_2026 .bn-card').forEach(function(o) {{
        var t = o.getAttribute('data-tipo');
        var c = o.getAttribute('data-clasif');
        var okTipo = tipoTodos || tiposActivos.indexOf(t) !== -1;
        var okClasif = clasifTodas || clasifsActivas.indexOf(c) !== -1;
        if (okTipo && okClasif) {{ o.classList.remove('oculta'); visibles++; }}
        else {{ o.classList.add('oculta'); }}
    }});
    var contador = document.getElementById('contador-ofertas');
    if (contador) contador.textContent = visibles;
}}
{JS_MARKER_FIN}
"""


# ---------------------------------------------------------------------------
# Reemplazos en el HTML
# ---------------------------------------------------------------------------

def _reemplazar_bloque_ofertas(html_actual, bloque_nuevo):
    """Reemplaza el contenido entre los marcadores si existen. Si no
    existen, reemplaza el contenido completo del `<div class="card">` que
    está dentro de `id="ofertas_2026"` desde después del <p> introductorio
    hasta antes del cierre `</div></div>` de la sección.

    Estrategia robusta: localiza el <h2 class="card-title">Ofertas 2026</h2>
    y el primer <p> que le sigue (intro), preserva ambos, y reemplaza todo
    lo que va desde después de ese <p> hasta el cierre `</div></div>` que
    cierra el content-section.
    """
    if MARKER_INI in html_actual and MARKER_FIN in html_actual:
        # Idempotente: reemplaza solo el bloque entre marcadores
        return re.sub(
            re.escape(MARKER_INI) + r".*?" + re.escape(MARKER_FIN),
            bloque_nuevo,
            html_actual,
            count=1,
            flags=re.DOTALL,
        )

    # Primera vez: localiza la sección ofertas_2026 y reemplaza todo
    # desde después del primer <p> introductorio hasta el cierre del
    # content-section. El lookahead exige que después vengan otra
    # content-section o </main>, NUNCA un simple comentario `<!--`
    # porque el HTML original tenía comentarios `<!-- Tipo: ... -->`
    # internos que confundían al regex y dejaban contenido residual.
    patron = re.compile(
        r'(<div class="content-section"[^>]*id="ofertas_2026"[^>]*>\s*<div class="card"[^>]*>\s*'
        r'(?:<img[^>]*>\s*)?'
        r'(?:<div[^>]*>\s*)?'
        r'<h2 class="card-title">Ofertas 2026</h2>\s*'
        r'(?:<p[^>]*>.*?</p>\s*)?)'
        r'.*?'
        r'(\s*(?:</div>\s*)?</div>\s*</div>\s*(?=<div class="content-section"|</main>))',
        re.DOTALL,
    )
    match = patron.search(html_actual)
    if not match:
        raise RuntimeError(
            "No se encontró la sección Ofertas 2026 en el HTML. "
            "Verifica que exista <div ... id=\"ofertas_2026\"> con <h2>Ofertas 2026</h2>."
        )
    return (
        html_actual[:match.end(1)]
        + "\n\n                " + bloque_nuevo + "\n"
        + html_actual[match.start(2):]
    )


def _inyectar_css(html_actual):
    if CSS_MARKER_INI in html_actual:
        return re.sub(
            re.escape(CSS_MARKER_INI) + r".*?" + re.escape(CSS_MARKER_FIN),
            CSS_CHIPS.strip(),
            html_actual,
            count=1,
            flags=re.DOTALL,
        )
    # Inyecta antes del </style>
    return html_actual.replace("</style>", CSS_CHIPS + "</style>", 1)


def _inyectar_js(html_actual):
    if JS_MARKER_INI in html_actual:
        return re.sub(
            re.escape(JS_MARKER_INI) + r".*?" + re.escape(JS_MARKER_FIN),
            JS_FILTROS.strip(),
            html_actual,
            count=1,
            flags=re.DOTALL,
        )
    # Inyecta antes del primer </script> que cierra el bloque principal
    # (el último </script> antes de </body>).
    idx = html_actual.rfind("</script>")
    if idx == -1:
        return html_actual + f"\n<script>{JS_FILTROS}</script>\n"
    return html_actual[:idx] + JS_FILTROS + html_actual[idx:]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generar():
    if not HTML_PATH.exists():
        raise RuntimeError(f"No existe {HTML_PATH}")
    html_actual = HTML_PATH.read_text(encoding="utf-8")
    ofertas = _leer_ofertas()
    mapeo_sirbe = _leer_mapeo_sirbe()

    bloque = _bloque_ofertas(ofertas, mapeo_sirbe)
    html_nuevo = _reemplazar_bloque_ofertas(html_actual, bloque)
    html_nuevo = _inyectar_css(html_nuevo)
    html_nuevo = _inyectar_js(html_nuevo)

    # Inyectar bloque de "Reporte a políticas" entre sus marcadores
    bloque_pol = _html_seccion_politicas()
    if MARKER_POL_INI in html_nuevo and MARKER_POL_FIN in html_nuevo:
        html_nuevo = re.sub(
            re.escape(MARKER_POL_INI) + r".*?" + re.escape(MARKER_POL_FIN),
            bloque_pol,
            html_nuevo,
            count=1,
            flags=re.DOTALL,
        )

    HTML_PATH.write_text(html_nuevo, encoding="utf-8")
    print(f"Actualizado: {HTML_PATH}")
    print(f"  Ofertas inyectadas: {len(ofertas)}")
    print(f"  Mapeos SIRBE encontrados: {len(mapeo_sirbe)}")
    def _hay_match(nombre, mapeo):
        n = _normalizar(nombre)
        if n in mapeo:
            return True
        for k in mapeo:
            if k and (k in n or n in k):
                return True
        return False

    coincidencias = sum(1 for o in ofertas if _hay_match(o['nombre'], mapeo_sirbe))
    print(f"  Ofertas con código SIRBE: {coincidencias}/{len(ofertas)}")


if __name__ == "__main__":
    generar()
