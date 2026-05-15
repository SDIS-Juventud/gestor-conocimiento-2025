# -*- coding: utf-8 -*-
"""Script único para generar el Excel inicial de mapeo Oferta → Política
del eje Bienestar. Se corre una sola vez. Luego Carolina edita el Excel
a mano y el script de generación del HTML lo lee.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

BASE = Path(__file__).resolve().parent.parent
RUTA = BASE / "ejes" / "Políticas" / "Mapeo_Politicas_bienestar.xlsx"

FILAS = [
    # (Nombre oferta, Tipo reporte, Tema, Código producto, Producto corto, Justificación)
    # Sin tabú (DSDR)
    ("Sin tabú, ni culpa: decidir es mi Derecho", "Política Pública", "Mujer y Equidad de Género", "6.1.9", "Sensibilización e información en DSDR", "Ciclo formativo en derechos sexuales y reproductivos."),
    ("Sin tabú, ni culpa: decidir es mi Derecho", "Política Pública", "Mujer y Equidad de Género", "1.2.13", "Sensibilización DSDR con enfoque de género", "El ciclo aborda DSDR con perspectiva de género."),
    ("Sin tabú, ni culpa: decidir es mi Derecho", "Política Pública", "Infancia y Adolescencia", "1.2.8", "Formación en DSDR a NNA y familias", "Dirigido a jóvenes en servicio social obligatorio (10° y 11°)."),

    # Parece normal (PVBG)
    ("Parece normal, pero no lo es. Identifica las red flags", "Política Pública", "Seguridad, paz y convivencia", "1.1.7", "Prevención de violencias y resolución de conflictos", "Aborda señales de alerta y prevención de VBG."),
    ("Parece normal, pero no lo es. Identifica las red flags", "Política Pública", "Mujer y Equidad de Género", "1.2.13", "Sensibilización DSDR con enfoque de género", "Trabaja red flags en relaciones desde enfoque de género."),

    # Brújula interior (SM)
    ("Brújula interior para conectarnos.", "Plan", "Conducta suicida", "Plan ICCS", "Espacios de fortalecimiento de capacidades", "Ciclo HSE / habilidades socioemocionales."),
    ("Brújula interior para conectarnos.", "Plan", "Conducta suicida", "Plan ICCS", "Espacios de reflexión sobre factores de riesgo en conducta suicida", "Trabaja factores de riesgo en salud mental."),

    # Bienestar 360 (mixto)
    ("Bienestar 360", "Política Pública", "Mujer y Equidad de Género", "6.1.9", "Sensibilización e información en DSDR", "Módulo DSDR del ciclo mixto."),
    ("Bienestar 360", "Política Pública", "Infancia y Adolescencia", "1.2.8", "Formación en DSDR a NNA y familias", "Módulo DSDR del ciclo mixto."),
    ("Bienestar 360", "Política Pública", "Seguridad, paz y convivencia", "1.1.7", "Prevención de violencias y resolución de conflictos", "Módulo PVBG del ciclo mixto."),
    ("Bienestar 360", "Plan", "Conducta suicida", "Plan ICCS", "Espacios de fortalecimiento de capacidades", "Módulo de salud mental del ciclo mixto."),

    # Parche Creativo (Experiencias)
    ("Parche Creativo pa' conectar", "Sistema", "Sistema Distrital de Formación Artística y Cultural (SIDFAC)", "SIDFAC", "Formación y experiencias en arte y cultura", "Las experiencias artísticas se reportan al sistema cultural."),
    ("Parche Creativo pa' conectar", "Política Pública", "Mujer y Equidad de Género", "1.2.13", "Sensibilización DSDR con enfoque de género", "Las experiencias abordan DSDR con enfoque de género."),

    # Píldoras
    ("En corto: Píldoras de Bienestar", "Política Pública", "Mujer y Equidad de Género", "6.1.9", "Sensibilización e información en DSDR", "Píldoras informativas en DSDR."),
    ("En corto: Píldoras de Bienestar", "Política Pública", "Seguridad, paz y convivencia", "1.1.7", "Prevención de violencias y resolución de conflictos", "Píldoras informativas en PVBG."),
    ("En corto: Píldoras de Bienestar", "Plan", "Conducta suicida", "Plan ICCS", "Espacios de reflexión sobre factores de riesgo en conducta suicida", "Píldoras informativas en salud mental."),

    # Bienestar con Rumbo (Salas de escucha)
    ("Bienestar con Rumbo", "Plan", "Comité Intersectorial de Salud", "Plan CIS", "Centros de escucha para la juventud", "Salas de escucha reportan directamente a este producto."),
    ("Bienestar con Rumbo", "Plan", "Conducta suicida", "Plan ICCS", "Espacios de fortalecimiento de capacidades", "Fortalecimiento de habilidades socioemocionales y orientación."),

    # Enrútate (Enrutamiento)
    ("Enrutate con tu Bienestar", "Plan", "Comité Intersectorial de Salud", "Plan CIS", "Orientación socio-ocupacional a jóvenes", "Enrutamiento individual a rutas de salud."),
    ("Enrutate con tu Bienestar", "Política Pública", "LGBTI", "1.2.12", "Orientación socio-ocupacional para jóvenes LGBTI", "Enrutamiento con enfoque diferencial LGBTI cuando aplica."),

    # Ruta al Derecho (Acceso)
    ("Ruta al Derecho", "Política Pública", "Mujer y Equidad de Género", "6.1.9", "Sensibilización e información en DSDR", "Brigadas de acceso a servicios DSDR."),
    ("Ruta al Derecho", "Política Pública", "LGBTI", "1.2.11", "Vinculación a oferta de prevención y DSDR para jóvenes LGBTI", "Brigadas con enfoque LGBTI cuando aplica."),
    ("Ruta al Derecho", "Política Pública", "Actividades Sexuales Pagadas", "2.2.7", "Jornadas de prevención y métodos anticonceptivos para jóvenes en ASP", "Brigadas con jóvenes en ASP cuando aplica."),

    # El Micrófono (Master class)
    ("El Micrófono es Nuestro: hablemos sin tabús", "Política Pública", "Mujer y Equidad de Género", "6.1.9", "Sensibilización e información en DSDR", "Foros conmemorativos sobre DSDR."),
    ("El Micrófono es Nuestro: hablemos sin tabús", "Política Pública", "Seguridad, paz y convivencia", "1.1.7", "Prevención de violencias y resolución de conflictos", "Foros conmemorativos sobre PVBG."),
    ("El Micrófono es Nuestro: hablemos sin tabús", "Plan", "Conducta suicida", "Plan ICCS", "Espacios de reflexión sobre factores de riesgo en conducta suicida", "Foros sobre salud mental."),

    # Tomar la voz
    ("Tomar la voz", "Política Pública", "Mujer y Equidad de Género", "6.1.9", "Sensibilización e información en DSDR", "Forma jóvenes multiplicadores en DSDR."),
    ("Tomar la voz", "Política Pública", "Infancia y Adolescencia", "1.2.8", "Formación en DSDR a NNA y familias", "Forma multiplicadores en sus territorios."),

    # Semana Andina (Evento)
    ("Semana Andina", "Política Pública", "Mujer y Equidad de Género", "6.1.9", "Sensibilización e información en DSDR", "Semana distrital DSDR."),
    ("Semana Andina", "Política Pública", "Mujer y Equidad de Género", "1.2.13", "Sensibilización DSDR con enfoque de género", "Semana distrital con perspectiva de género."),
    ("Semana Andina", "Política Pública", "Infancia y Adolescencia", "1.2.8", "Formación en DSDR a NNA y familias", "Semana de prevención de maternidad y paternidad temprana."),
    ("Semana Andina", "Política Pública", "LGBTI", "1.2.11", "Vinculación a oferta de prevención y DSDR para jóvenes LGBTI", "Semana con enfoque diferencial LGBTI cuando aplica."),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mapeo"

HEADERS = ["Nombre oferta", "Tipo de reporte", "Tema / Política", "Código producto", "Producto (corto)", "Justificación"]
ws.append(HEADERS)

header_fill = PatternFill(start_color="253C5C", end_color="253C5C", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
thin = Side(border_style="thin", color="DDDDDD")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

for col_idx in range(1, len(HEADERS) + 1):
    cell = ws.cell(1, col_idx)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = border

for fila in FILAS:
    ws.append(list(fila))

for r in range(2, len(FILAS) + 2):
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(r, c)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border
        cell.font = Font(size=10, name="Calibri")

anchos = [38, 18, 38, 20, 42, 50]
for i, w in enumerate(anchos, 1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"

RUTA.parent.mkdir(parents=True, exist_ok=True)
wb.save(RUTA)
print(f"Generado: {RUTA}")
print(f"  Filas: {len(FILAS)}")
print(f"  Ofertas con mapeo: {len(set(f[0] for f in FILAS))}")
print()
print("Ofertas SIN match en Casas (intencionalmente):")
print("  - Parche Consciente (SPA — no hay producto específico de SPA en Casas)")
print("  - Feria de Emprendimientos con Propósito")
print("  - Actívate: laboratorios de creación (estrategia interna para equipos)")
