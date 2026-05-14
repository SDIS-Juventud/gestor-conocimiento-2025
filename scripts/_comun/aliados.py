# -*- coding: utf-8 -*-
"""
Datos y funciones para renderizar la seccion de aliados de los gestores.

Fuente de verdad: los dos Excel en ../alianzas/
    - Matriz de seguimiento Alianzas Estrategicas (2025).xlsx, hoja '2025'
    - Matriz de seguimiento Alianzas Estrategicas (2026).xlsx, hoja '2026'

Cada vez que se regenera un gestor (generar_gc_jco.py, generar_gc_forjar.py)
se leen los Excel y se arma la seccion. Cualquier cambio en los Excel queda
reflejado automaticamente en el HTML. No hay contenido hardcodeado.

Formato de tarjetas (identico para 2025 y 2026):
    - Una tarjeta por aliado (aunque tenga varios programas).
    - Dentro: nombre del aliado (negrita) + lista de programas, cada uno
      con su nombre (cursiva) y su objetivo general.

Se filtran solo las alianzas con etapa "Ejecucion" o "Permanente" — las
demas (Negociacion, Planeacion, No iniciado, Finalizado, vacias) se omiten.
"""

import html
from pathlib import Path

import openpyxl
import pandas as pd


# --------------------------------------------------------------------
# Rutas de los Excel
# --------------------------------------------------------------------
_ALIANZAS_DIR = Path(__file__).resolve().parents[2] / "alianzas"
_EXCEL_2025 = _ALIANZAS_DIR / "Matriz de seguimiento Alianzas Estratégicas (2025).xlsx"
_EXCEL_2026 = _ALIANZAS_DIR / "Matriz de seguimiento Alianzas Estratégicas (2026).xlsx"

# Excel con los enlaces externos que se renderizan como botones en el HTML.
# Estructura: HTML | SECCION | ENLACE. La columna SECCION se usa como clave
# (p. ej. "Alianzas 2025", "Alianzas 2026"). Editar este Excel basta para
# cambiar los URLs en los gestores; no hay que tocar codigo Python.
_ENLACES_PATH = Path(__file__).resolve().parents[2] / "enlaces" / "enlaces.xlsx"


def _cargar_enlaces():
    """Lee enlaces/enlaces.xlsx y devuelve {seccion: url}.

    Si el archivo no existe o no tiene filas validas, devuelve un dict vacio
    y los botones simplemente no se renderizan (los gestores siguen funcionando).

    Busca la hoja cuyas columnas A/B/C sean HTML/SECCION/ENLACE (sin asumir
    que sea la primera ni la activa, porque el Excel se editó y la hoja
    activa quedó apuntando a otra cosa). De fallback usa 'Hoja1'.
    """
    if not _ENLACES_PATH.exists():
        return {}
    wb = openpyxl.load_workbook(_ENLACES_PATH, data_only=True)
    ws = None
    for nombre in wb.sheetnames:
        prueba = wb[nombre]
        h1 = str(prueba.cell(1, 1).value or "").strip().upper()
        h2 = str(prueba.cell(1, 2).value or "").strip().upper()
        h3 = str(prueba.cell(1, 3).value or "").strip().upper()
        if h1 == "HTML" and h2 == "SECCION" and h3 == "ENLACE":
            ws = prueba
            break
    if ws is None:
        ws = wb["Hoja1"] if "Hoja1" in wb.sheetnames else wb.active
    enlaces = {}
    for r in range(2, ws.max_row + 1):
        seccion = ws.cell(r, 2).value
        url = ws.cell(r, 3).value
        if seccion and url:
            enlaces[str(seccion).strip()] = str(url).strip()
    return enlaces


# --------------------------------------------------------------------
# Helpers de limpieza y clasificacion
# --------------------------------------------------------------------

def _limpiar(valor):
    """Normaliza un texto leido del Excel para inyectarlo como HTML.

    Hace trim, colapsa espacios multiples y saltos de linea en un solo
    espacio, y escapa caracteres especiales de HTML (<, >, &, ', ").
    Devuelve '' si el valor es NaN o vacio.
    """
    if pd.isna(valor):
        return ""
    texto = str(valor).strip()
    # Colapsa cualquier secuencia de espacios o saltos de linea
    texto = " ".join(texto.split())
    return html.escape(texto)


def _clasificar_servicio(valor):
    """Mapea el valor de la columna 'Servicio' a claves estandar.

    Devuelve un conjunto con alguna combinacion de 'cdj', 'jco', 'forjar'.
    Soporta variantes como 'CDJ', 'Casas de Juventud', 'JCO',
    'Jovenes con oportunidades', 'FORJAR', 'Forjar Restaurativo', o valores
    multilinea como 'Casas de Juventud\\nForjar Restaurativo'.
    """
    if pd.isna(valor):
        return set()
    txt = str(valor).upper()
    # Quita tildes basicas para matching robusto
    for tilde, plano in [('Á', 'A'), ('É', 'E'), ('Í', 'I'), ('Ó', 'O'), ('Ú', 'U')]:
        txt = txt.replace(tilde, plano)
    claves = set()
    if 'CDJ' in txt or 'CASAS' in txt:
        claves.add('cdj')
    if 'JCO' in txt or 'JOVENES' in txt:
        claves.add('jco')
    if 'FORJAR' in txt:
        claves.add('forjar')
    return claves


# Etapas que se consideran alianzas mostrables en la web.
# Incluye Finalización porque las alianzas que cerraron en 2025 igual
# estuvieron vigentes ese año y deben quedar reflejadas.
_ETAPAS_VALIDAS = {'EJECUCION', 'PERMANENTE', 'FINALIZACION'}


def _etapa_valida(valor):
    """Decide si una alianza debe mostrarse segun su etapa."""
    if pd.isna(valor):
        return False
    txt = str(valor).upper().strip()
    for tilde, plano in [('Á', 'A'), ('É', 'E'), ('Í', 'I'), ('Ó', 'O'), ('Ú', 'U')]:
        txt = txt.replace(tilde, plano)
    return txt in _ETAPAS_VALIDAS


# --------------------------------------------------------------------
# Carga de datos
# --------------------------------------------------------------------

def _leer_excel_con_merged(path, hoja, skiprows):
    """Lee el Excel respetando celdas fusionadas (merged cells).

    Para cada rango fusionado propaga el valor del extremo superior
    izquierdo a todas las celdas del rango, replicando lo que el usuario
    ve en Excel. Esto es necesario porque en la matriz de alianzas 2025
    varios aliados ocupan varias filas y solo la primera tiene el nombre,
    así que pandas las descartaba.

    Devuelve un DataFrame con las filas a partir de skiprows+1 (1-indexed),
    sin encabezados.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[hoja]
    # Mapa de celdas fusionadas: (fila, col) -> valor top-left
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(row=mr.min_row, column=mr.min_col).value
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merge_map[(row, col)] = top_val
    # Construir las filas a partir de skiprows+1 (1-indexed)
    filas = []
    for r in range(skiprows + 1, ws.max_row + 1):
        fila = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                v = merge_map.get((r, c))
            fila.append(v)
        filas.append(fila)
    return pd.DataFrame(filas)


def _cargar_excel(path, hoja, col_servicio, col_aliado, col_proyecto,
                  col_objetivo, col_etapa):
    """Lee un Excel de alianzas, filtra por etapa y agrupa por servicio+aliado.

    Devuelve {clave_servicio: [(aliado, [(proyecto, descripcion), ...]), ...]}
    donde cada aliado aparece una sola vez con todos sus programas.
    Solo se incluyen filas con etapa en _ETAPAS_VALIDAS.
    """
    df = _leer_excel_con_merged(path, hoja, skiprows=4)
    # Estructura temporal: {servicio: {aliado: [(proyecto, desc), ...]}}
    temp = {'cdj': {}, 'jco': {}, 'forjar': {}}
    for _, row in df.iterrows():
        aliado = row.iloc[col_aliado] if len(row) > col_aliado else None
        if pd.isna(aliado) or str(aliado).strip() == "":
            continue
        etapa = row.iloc[col_etapa] if len(row) > col_etapa else None
        if not _etapa_valida(etapa):
            continue
        servicio_raw = row.iloc[col_servicio] if len(row) > col_servicio else None
        proyecto = row.iloc[col_proyecto] if len(row) > col_proyecto else None
        descripcion = row.iloc[col_objetivo] if len(row) > col_objetivo else None
        nombre = _limpiar(aliado)
        prog = (_limpiar(proyecto), _limpiar(descripcion))
        for clave in _clasificar_servicio(servicio_raw):
            progs = temp[clave].setdefault(nombre, [])
            # Evita duplicados: al resolver celdas fusionadas, el mismo
            # programa puede aparecer en varias filas de cronograma.
            if prog not in progs:
                progs.append(prog)

    # Convierte a lista ordenada: [(aliado, [programas]), ...]
    por_servicio = {}
    for clave, aliados_dict in temp.items():
        aliados_ordenados = sorted(aliados_dict.items(), key=lambda t: t[0].lower())
        por_servicio[clave] = aliados_ordenados
    return por_servicio


def _cargar_2025():
    """Excel 2025, hoja '2025'. Cols: 0=Aliado, 4=Servicio, 6=Proyecto, 7=Objetivo, 13=Etapa."""
    return _cargar_excel(
        path=_EXCEL_2025, hoja='2025',
        col_servicio=4, col_aliado=0, col_proyecto=6, col_objetivo=7, col_etapa=13,
    )


def _cargar_2026():
    """Excel 2026, hoja '2026'. Cols: 0=Servicio, 1=Aliado, 3=Proyecto, 11=Objetivo, 4=Etapa."""
    return _cargar_excel(
        path=_EXCEL_2026, hoja='2026',
        col_servicio=0, col_aliado=1, col_proyecto=3, col_objetivo=11, col_etapa=4,
    )


# --------------------------------------------------------------------
# Renderizadores de tarjetas
# --------------------------------------------------------------------

def _bloque_programa(proyecto, descripcion, con_separador):
    """Un programa dentro de la tarjeta: proyecto en cursiva + descripcion."""
    estilo_sep = (
        ' margin-top:12px; padding-top:10px; border-top:1px dashed #e0e0e0;'
        if con_separador else ' margin-top:8px;'
    )
    lineas = f'<div style="{estilo_sep}">'
    if proyecto:
        lineas += (
            f'<p style="font-size:0.85rem; color:#555; margin:0;">'
            f'<em>{proyecto}</em></p>'
        )
    if descripcion:
        lineas += (
            f'<p style="font-size:0.85rem; color:#666; margin:3px 0 0;">'
            f'{descripcion}</p>'
        )
    lineas += '</div>'
    return lineas


def _card(aliado, programas):
    """Tarjeta con un aliado y todos sus programas.

    programas: lista de tuplas (proyecto, descripcion).
    """
    bloques = "\n                        ".join(
        _bloque_programa(p, d, con_separador=(i > 0))
        for i, (p, d) in enumerate(programas)
    )
    return (
        '                    <div class="card" style="margin-bottom:10px; '
        'padding:15px 20px;">\n'
        f'                        <strong style="color:#3A3A3A;">{aliado}</strong>\n'
        f'                        {bloques}\n'
        '                    </div>'
    )


# --------------------------------------------------------------------
# Secciones completas por servicio
# --------------------------------------------------------------------

_SUBTITULO_ANIO_STYLE = (
    "font-size:1.1rem; color:var(--accent); margin-top:25px; "
    "margin-bottom:15px; padding-top:20px; "
    "border-top:1px solid #e0e0e0;"
)


def _bloque_anio(titulo, cards, es_primero, link=None, color_acento="var(--accent)"):
    """Devuelve el HTML del titulo + tarjetas para un anio. '' si no hay cards.

    Si un anio no tiene aliados (p.ej. Forjar 2025 tras aplicar el filtro),
    se omite por completo el bloque: ni titulo ni contenedor.

    Si se pasa `link`, se agrega al final un boton "Ver matriz completa"
    con esa URL (apuntando al SharePoint con la matriz oficial). El color
    del borde y del texto del boton se controla con `color_acento`: por
    defecto usa la variable CSS `--accent` (JCO, Forjar). Casas de
    Juventud no define esa variable, asi que pasa su color institucional
    `#253C5C` directamente.
    """
    if not cards:
        return ""
    extras = " margin-top:0; padding-top:0; border-top:none;" if es_primero else ""
    boton = ""
    if link:
        boton = (
            '\n                    <div style="margin-top:14px;">\n'
            f'                        <a href="{link}" target="_blank" rel="noopener" '
            f'style="display:inline-block; padding:7px 16px; background:#ffffff; '
            f'border:1px solid {color_acento}; border-radius:6px; color:{color_acento}; '
            'text-decoration:none; font-size:0.85rem; font-weight:600;">Ver matriz completa &rarr;</a>\n'
            '                    </div>'
        )
    return (
        f'                    <h3 class="card-subtitle" style="{_SUBTITULO_ANIO_STYLE}{extras}">{titulo}</h3>\n'
        f'                    <div style="padding-left:0;">\n'
        f'{cards}\n'
        f'                    </div>{boton}'
    )


def _seccion(clave_servicio, nombre_servicio, color_acento="var(--accent)"):
    """HTML completo de la seccion de aliados para un servicio."""
    datos_2025 = _cargar_2025().get(clave_servicio, [])
    datos_2026 = _cargar_2026().get(clave_servicio, [])
    cards_2025 = "\n".join([_card(aliado, progs) for aliado, progs in datos_2025])
    cards_2026 = "\n".join([_card(aliado, progs) for aliado, progs in datos_2026])

    # URLs externas se leen del Excel enlaces/enlaces.xlsx en cada render.
    enlaces = _cargar_enlaces()
    link_2025 = enlaces.get("Alianzas 2025")
    link_2026 = enlaces.get("Alianzas 2026")

    # Si un anio queda vacio se omite del HTML (ni titulo ni contenedor).
    # El primer bloque que si tenga contenido pierde el separador superior.
    bloques = []
    if cards_2025:
        bloques.append(_bloque_anio("Alianzas 2025", cards_2025, es_primero=True, link=link_2025, color_acento=color_acento))
        if cards_2026:
            bloques.append(_bloque_anio("Alianzas 2026", cards_2026, es_primero=False, link=link_2026, color_acento=color_acento))
    elif cards_2026:
        bloques.append(_bloque_anio("Alianzas 2026", cards_2026, es_primero=True, link=link_2026, color_acento=color_acento))

    if not bloques:
        contenido = (
            '                    <p style="color:#999; font-style:italic;">'
            'Sin alianzas registradas.</p>'
        )
    else:
        contenido = "\n\n".join(bloques)

    return f'''            <div class="content-section" id="aliados">
                <div class="card">
                    <h2 class="card-title">Aliados</h2>
                    <p style="color:#666; margin-bottom:20px;">Entidades y organizaciones que apoyan la operaci&oacute;n del servicio {nombre_servicio}.</p>

{contenido}
                </div>
            </div>'''


def seccion_jco():
    """Seccion de aliados para Jovenes con Oportunidades."""
    return _seccion('jco', 'J&oacute;venes con Oportunidades')


def seccion_forjar():
    """Seccion de aliados para Forjar Restaurativo."""
    return _seccion('forjar', 'Forjar Restaurativo')


def seccion_casas():
    """Seccion de aliados para Casas de Juventud.

    El HTML de Casas no usa la variable CSS --accent (tiene los colores
    hardcodeados), asi que el boton recibe el azul institucional directo.
    """
    return _seccion('cdj', 'Casas de Juventud', color_acento='#253C5C')
